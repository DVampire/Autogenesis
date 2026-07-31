import openpyxl
from openpyxl.utils import get_column_letter
from typing import Any, Dict, Optional

from pydantic import ConfigDict, Field

from autogenesis.environment.server import environment_manager
from autogenesis.environment.types import Environment
from autogenesis.logger import logger
from autogenesis.registry import ENVIRONMENT

@ENVIRONMENT.register_module(force=True)
class GridMazeEnvironment(Environment):
    """Grid maze loaded from an Excel map."""

    model_config = ConfigDict(arbitrary_types_allowed=True, extra="allow")

    name: str = Field(default="grid_maze_environment")
    description: str = Field(default="Grid maze loaded from an Excel map.")
    metadata: Dict[str, Any] = Field(default_factory=dict)
    enable_evolving: bool = Field(default=True)

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.grid = {}
        self.pos = None
        self.came_from = None
        self.turn = 0
        self.start_pos = None
        self.end_pos = None
        self.history = []

    @environment_manager.action(
        name="load_map",
        description="Load grid maze from an Excel map. Args: path (str)."
    )
    async def load_map(self, path: str, **kwargs) -> Dict[str, Any]:
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            self.grid.clear()
            self.start_pos = None
            self.end_pos = None
            self.history = []
            
            for row in range(1, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    
                    if hasattr(cell.fill, 'start_color') and hasattr(cell.fill.start_color, 'rgb'):
                        rgb = cell.fill.start_color.rgb
                        if isinstance(rgb, str):
                            self.grid[(row, col)] = rgb
                            
                    if cell.value == "START":
                        self.start_pos = (row, col)
                    elif cell.value == "END":
                        self.end_pos = (row, col)
                        
            return {"success": True, "message": "Map loaded successfully."}
        except Exception as e:
            return {"success": False, "error": str(e)}

    @environment_manager.action(
        name="reset",
        description="Reset the environment to the START cell."
    )
    async def reset(self, **kwargs) -> Dict[str, Any]:
        if not self.start_pos:
            return {"success": False, "error": "No START cell found."}
        self.pos = self.start_pos
        self.came_from = None
        self.turn = 0
        self.history = []
        return {"success": True, "message": "Reset to START."}

    @environment_manager.action(
        name="move",
        description="Move up, down, left, or right. Args: direction (str)."
    )
    async def move(self, direction: str, **kwargs) -> Dict[str, Any]:
        if not self.pos:
            return {"success": False, "error": "Game not started."}
            
        row, col = self.pos
        
        dr, dc = 0, 0
        reverse_dir = ""
        if direction == "up":
            dr, dc = -1, 0
            reverse_dir = "down"
        elif direction == "down":
            dr, dc = 1, 0
            reverse_dir = "up"
        elif direction == "left":
            dr, dc = 0, -1
            reverse_dir = "right"
        elif direction == "right":
            dr, dc = 0, 1
            reverse_dir = "left"
        else:
            return {"success": False, "error": f"Invalid direction: {direction}"}

        if self.came_from == reverse_dir:
            return {"success": False, "error": "Cannot move reverse of the previous turn."}

        c1 = (row + dr, col + dc)

        if not (1 <= c1[0] <= 20 and 1 <= c1[1] <= 9):
            return {"success": False, "error": "First step is off-grid."}
        if self.grid.get(c1) == "FF0099FF":
            return {"success": False, "error": "First step is blue."}

        valid_second_steps = []
        for d, dr2, dc2, rev_d in [("up", -1, 0, "down"), ("down", 1, 0, "up"), ("left", 0, -1, "right"), ("right", 0, 1, "left")]:
            c2 = (c1[0] + dr2, c1[1] + dc2)
            if c2 == self.pos:
                continue
            if 1 <= c2[0] <= 20 and 1 <= c2[1] <= 9:
                if self.grid.get(c2) != "FF0099FF":
                    valid_second_steps.append((c2, d))
                    
        if not valid_second_steps:
            return {"success": False, "error": "No valid second step found."}
            
        c2, second_dir = valid_second_steps[0]

        self.history.append((self.pos, self.came_from, self.turn))
        
        self.pos = c2
        self.came_from = second_dir
        self.turn += 1

        return {"success": True, "message": f"Moved {direction}."}

    @environment_manager.action(
        name="undo",
        description="Take back the last step, restoring position and came-from."
    )
    async def undo(self, **kwargs) -> Dict[str, Any]:
        if not self.history:
            return {"success": False, "error": "No moves to undo."}
        
        self.pos, self.came_from, self.turn = self.history.pop()
        return {"success": True, "message": "Undo successful."}

    def get_state(self) -> Dict[str, Any]:
        if not self.pos:
            return {"error": "Game not started."}
            
        row, col = self.pos
        a1_notation = f"{get_column_letter(col)}{row}"
        current_color = self.grid.get(self.pos)
        
        def simulate_turn(p, cf, direction):
            r, c = p
            dr, dc = 0, 0
            reverse_dir = ""
            if direction == "up":
                dr, dc = -1, 0
                reverse_dir = "down"
            elif direction == "down":
                dr, dc = 1, 0
                reverse_dir = "up"
            elif direction == "left":
                dr, dc = 0, -1
                reverse_dir = "right"
            elif direction == "right":
                dr, dc = 0, 1
                reverse_dir = "left"

            if cf == reverse_dir:
                return None

            c1 = (r + dr, c + dc)
            if not (1 <= c1[0] <= 20 and 1 <= c1[1] <= 9):
                return None
            if self.grid.get(c1) == "FF0099FF":
                return None

            for d2, dr2, dc2, rev_d in [("up", -1, 0, "down"), ("down", 1, 0, "up"), ("left", 0, -1, "right"), ("right", 0, 1, "left")]:
                c2 = (c1[0] + dr2, c1[1] + dc2)
                if c2 == p:
                    continue
                if 1 <= c2[0] <= 20 and 1 <= c2[1] <= 9:
                    if self.grid.get(c2) != "FF0099FF":
                        return c2, d2
            return None

        def can_survive(p, cf, turns_remaining):
            if p == self.end_pos:
                return True
            if turns_remaining == 0:
                return True
            for d in ["up", "down", "left", "right"]:
                res = simulate_turn(p, cf, d)
                if res:
                    np, ncf = res
                    if can_survive(np, ncf, turns_remaining - 1):
                        return True
            return False

        legal_moves = []
        for d in ["up", "down", "left", "right"]:
            res = simulate_turn(self.pos, self.came_from, d)
            if res:
                np, ncf = res
                if can_survive(np, ncf, 3):
                    legal_moves.append(d)
                    
        stuck = len(legal_moves) == 0
                    
        return {
            "pos": a1_notation,
            "turn": self.turn,
            "color": current_color,
            "stuck": stuck,
            "legal_moves": legal_moves
        }
